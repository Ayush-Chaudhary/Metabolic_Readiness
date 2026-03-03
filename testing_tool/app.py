"""
Metabolic Readiness — Clinical Testing Tool
=============================================
Streamlit front-end for the clinical team to test insight generation
with synthetic patient data.  All interactions are logged to a local
Excel file for traceability.
"""

import streamlit as st
import pandas as pd
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

from config import connection_status, get_databricks_host, get_databricks_token, has_oauth_m2m, has_llm_sp_creds
from backend import (
    generate_synthetic_context,
    run_pipeline,
    get_feature_snapshot,
    GLUCOSE_SCENARIOS,
    STEP_SCENARIOS,
    ACTIVITY_SCENARIOS,
    SLEEP_SCENARIOS,
    FOOD_SCENARIOS,
    MED_SCENARIOS,
    WEIGHT_SCENARIOS,
    MENTAL_WELLBEING_SCENARIOS,
    JOURNEY_SCENARIOS,
    EXERCISE_VIDEO_SCENARIOS,
    EXERCISE_PROGRAM_SCENARIOS,
)

# =============================================================================
# PAGE CONFIG
# =============================================================================
st.set_page_config(
    page_title="SIMON Metabolic Readiness — Testing Tool",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded",
)

LOG_FILE = os.path.join(os.path.dirname(__file__), "test_log.xlsx")


# =============================================================================
# EXCEL LOGGING
# =============================================================================

def _ensure_log_file():
    """Create the Excel log file with headers if it doesn't exist."""
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Log"
        headers = [
            "Timestamp", "Tester Name",
            # Config
            "A1C Target Group", "User Focus", "Weight Goal Type",
            "Has CGM", "Has Step Tracker", "Has Medications",
            "Glucose Scenario", "Step Scenario", "Activity Scenario",
            "Sleep Scenario", "Food Scenario", "Med Scenario",
            "Weight Scenario", "Mental Wellbeing", "Journey Scenario",
            "Exercise Video", "Exercise Program",
            # Model params
            "Temperature", "Max Tokens", "Max Positive Actions", "Greeting",
            # Results
            "Patient ID", "Daily Rating", "Score Description",
            "Positive Actions Used", "Opportunity Used",
            "Generated Message", "Character Count", "Word Count",
            "LLM Success",
            # Feedback
            "Quality Rating (1-5)", "Clinically Accurate",
            "Appropriate Tone", "Actionable Suggestion",
            "Feedback Notes",
        ]
        ws.append(headers)
        # Freeze header row and bold headers
        ws.freeze_panes = "A2"
        from openpyxl.styles import Font
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
        wb.save(LOG_FILE)


def _append_log_row(row_data: list):
    """Append a row to the Excel log."""
    _ensure_log_file()
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    ws.append(row_data)
    wb.save(LOG_FILE)


# =============================================================================
# SIDEBAR — Configuration Panel
# =============================================================================

st.sidebar.title("🧪 Test Configuration")
st.sidebar.markdown("---")

tester_name = st.sidebar.text_input("Your Name (optional)", placeholder="e.g., Dr. Smith")

st.sidebar.subheader("👤 Patient Profile")
col1, col2 = st.sidebar.columns(2)
with col1:
    a1c_group = st.selectbox(
        "A1C Target Group",
        ["DM Target <7%", "DM Target <8%", "DIP (Diabetes in Pregnancy)", "Non-DM"],
        index=0,
    )
with col2:
    user_focus = st.selectbox(
        "User Focus Area",
        ["None", "Weight", "Glucose", "Activity", "Eating Habits", "Sleep", "Medications", "Anxiety"],
        index=0,
    )

col3, col4 = st.sidebar.columns(2)
with col3:
    weight_goal = st.selectbox("Weight Goal", ["None", "Lose", "Maintain", "Gain"], index=0)
with col4:
    greeting_override = st.selectbox("Greeting", ["Auto", "Morning", "Afternoon", "Evening"], index=0)

st.sidebar.subheader("📱 Device / Feature Flags")
dcol1, dcol2, dcol3 = st.sidebar.columns(3)
with dcol1:
    has_cgm = st.checkbox("Has CGM", value=True)
with dcol2:
    has_step_tracker = st.checkbox("Step Tracker", value=True)
with dcol3:
    has_medications = st.checkbox("Medications", value=True)

st.sidebar.markdown("---")
st.sidebar.subheader("📊 Health Metric Scenarios")

glucose_scenario = st.sidebar.selectbox("Glucose Performance", list(GLUCOSE_SCENARIOS.keys()), index=1)
step_scenario = st.sidebar.selectbox("Daily Steps", list(STEP_SCENARIOS.keys()), index=1)
activity_scenario = st.sidebar.selectbox("Weekly Activity", list(ACTIVITY_SCENARIOS.keys()), index=1)
sleep_scenario = st.sidebar.selectbox("Sleep Quality", list(SLEEP_SCENARIOS.keys()), index=1)
food_scenario = st.sidebar.selectbox("Food Logging", list(FOOD_SCENARIOS.keys()), index=1)
med_scenario = st.sidebar.selectbox("Medication Adherence", list(MED_SCENARIOS.keys()), index=1)
weight_scenario = st.sidebar.selectbox("Weight Trend", list(WEIGHT_SCENARIOS.keys()), index=1)
mental_scenario = st.sidebar.selectbox("Mental Wellbeing", list(MENTAL_WELLBEING_SCENARIOS.keys()), index=1)

st.sidebar.markdown("---")
st.sidebar.subheader("🎯 Engagement Features")

journey_scenario = st.sidebar.selectbox("Journey", list(JOURNEY_SCENARIOS.keys()), index=3)
exercise_video_scenario = st.sidebar.selectbox("Exercise Video", list(EXERCISE_VIDEO_SCENARIOS.keys()), index=3)
exercise_program_scenario = st.sidebar.selectbox("Exercise Program", list(EXERCISE_PROGRAM_SCENARIOS.keys()), index=3)

st.sidebar.markdown("---")
st.sidebar.subheader("⚙️ Model Parameters")

temperature = st.sidebar.slider("LLM Temperature", 0.0, 1.0, 0.7, 0.05)
max_tokens = st.sidebar.slider("Max Tokens", 100, 500, 180, 10)
max_actions = st.sidebar.selectbox("Max Positive Actions", [1, 2, 3], index=1)

# --- Connection status (credentials come from env vars / .env, not user input) ---
_conn = connection_status()
st.sidebar.markdown("---")
if _conn["configured"]:
    st.sidebar.success(_conn["label"])
else:
    st.sidebar.warning(_conn["label"])

st.sidebar.markdown("---")
generate_btn = st.sidebar.button("🚀 Generate Insight", type="primary", use_container_width=True)

# Download log button
if os.path.exists(LOG_FILE):
    with open(LOG_FILE, "rb") as f:
        st.sidebar.download_button(
            "📥 Download Test Log",
            f,
            file_name="metabolic_readiness_test_log.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# =============================================================================
# MAIN AREA
# =============================================================================

st.title("SIMON Metabolic Readiness — Insight Testing Tool")
st.markdown(
    "Configure a synthetic patient profile in the sidebar, generate an insight, "
    "review the results, and submit feedback. All interactions are logged to an Excel file."
)

# Initialize session state
if "result" not in st.session_state:
    st.session_state.result = None
if "user_context" not in st.session_state:
    st.session_state.user_context = None
if "logged" not in st.session_state:
    st.session_state.logged = False

# =============================================================================
# GENERATE
# =============================================================================

if generate_btn:
    with st.spinner("Generating synthetic patient data and running pipeline..."):
        # Build synthetic context
        user_ctx = generate_synthetic_context(
            a1c_target_group=a1c_group,
            user_focus=user_focus,
            weight_goal_type=weight_goal,
            has_cgm=has_cgm,
            has_step_tracker=has_step_tracker,
            has_medications=has_medications,
            glucose_scenario=glucose_scenario,
            step_scenario=step_scenario,
            activity_scenario=activity_scenario,
            sleep_scenario=sleep_scenario,
            food_scenario=food_scenario,
            med_scenario=med_scenario,
            weight_scenario=weight_scenario,
            mental_scenario=mental_scenario,
            journey_scenario=journey_scenario,
            exercise_video_scenario=exercise_video_scenario,
            exercise_program_scenario=exercise_program_scenario,
        )

        # Run pipeline (credentials loaded from env vars / .env via config.py)
        result = run_pipeline(
            user_context=user_ctx,
            temperature=temperature,
            max_tokens=max_tokens,
            max_positive_actions=max_actions,
            greeting_override=greeting_override,
            databricks_host=get_databricks_host(),
            databricks_token=get_databricks_token(),
        )

        st.session_state.result = result
        st.session_state.user_context = user_ctx
        st.session_state.logged = False

# =============================================================================
# DISPLAY RESULTS
# =============================================================================

result = st.session_state.result
user_ctx = st.session_state.user_context

if result is not None:
    st.markdown("---")

    # --- Message Card ---
    rating = result["rating"]
    rating_emoji = {
        "Committed": "🟢",
        "Strong": "🔵",
        "Consistent": "🟣",
        "Building": "🟠",
        "Ready": "⚪",
    }
    if result["success"]:
        llm_status = "🟢 Real LLM (Databricks)"
    elif has_llm_sp_creds() or has_oauth_m2m() or get_databricks_token():
        llm_status = "❌ LLM call failed — check endpoint permissions (see README for fix)"
    else:
        llm_status = "⚠️ Mock response — configure .env with DATABRICKS credentials for real output"

    with st.container(border=True):
        meta_col, stat_col = st.columns([3, 2])
        with meta_col:
            st.markdown(
                f"**{rating_emoji.get(rating, '')} Daily Rating: {rating}** — "
                f"*{result['rating_description']}*"
            )
        with stat_col:
            st.caption(
                f"{result['character_count']} chars · {result['word_count']} words · {llm_status}"
            )
        st.markdown("---")
        st.markdown(result["message"])

    # --- Selected Content ---
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("#### ✅ Positive Actions Used")
        for action in result["positive_actions_used"]:
            st.markdown(
                f"- **`{action['key']}`** ({action['category']}): _{action['text']}_"
            )

    with col_b:
        st.markdown("#### 💡 Opportunity Used")
        opp = result["opportunity_used"]
        st.markdown(f"- **`{opp['key']}`** ({opp['category']}): _{opp['text']}_")

    # --- Score Breakdown ---
    with st.expander("📊 Score Breakdown & Rating", expanded=False):
        st.markdown(f"**Daily Rating:** {result['rating']}")
        st.markdown(f"**Description:** {result['rating_description']}")
        st.info(
            "Rating tiers: Committed (81-100%) · Strong (61-80%) · "
            "Consistent (41-60%) · Building (21-40%) · Ready (1-20%)"
        )

    # --- All Eligible Content ---
    with st.expander(f"📋 All Eligible Actions ({len(result['all_eligible_actions'])} found)", expanded=False):
        if result["all_eligible_actions"]:
            actions_df = pd.DataFrame(result["all_eligible_actions"])
            # Serialize dict/list columns to strings to avoid Arrow type issues
            for col in actions_df.columns:
                if actions_df[col].dtype == object:
                    actions_df[col] = actions_df[col].apply(
                        lambda x: json.dumps(x) if isinstance(x, (dict, list)) else str(x) if x is not None else ""
                    )
            st.dataframe(actions_df, use_container_width=True)
        else:
            st.warning("No eligible actions found.")

    with st.expander(f"📋 All Eligible Opportunities ({len(result['all_eligible_opportunities'])} found)", expanded=False):
        if result["all_eligible_opportunities"]:
            opps_df = pd.DataFrame(result["all_eligible_opportunities"])
            for col in opps_df.columns:
                if opps_df[col].dtype == object:
                    opps_df[col] = opps_df[col].apply(
                        lambda x: json.dumps(x) if isinstance(x, (dict, list)) else str(x) if x is not None else ""
                    )
            st.dataframe(opps_df, use_container_width=True)
        else:
            st.warning("No eligible opportunities found.")

    # --- Patient Feature Snapshot ---
    with st.expander("🔬 Synthetic Patient Features", expanded=False):
        snapshot = get_feature_snapshot(user_ctx)
        snap_df = pd.DataFrame(
            [{"Feature": k, "Value": str(v) if v is not None else "—"} for k, v in snapshot.items()]
        )
        st.dataframe(snap_df, use_container_width=True, height=500)

    # --- Prompt Inspector ---
    with st.expander("🔍 Prompt Inspector (System + User Prompts)", expanded=False):
        st.markdown("**System Prompt:**")
        st.code(result["system_prompt"], language="text")
        st.markdown("**User Prompt:**")
        st.code(result["user_prompt"], language="text")

    # ==========================================================================
    # FEEDBACK SECTION
    # ==========================================================================
    st.markdown("---")
    st.subheader("📝 Tester Feedback")

    fcol1, fcol2 = st.columns(2)
    with fcol1:
        quality_rating = st.select_slider(
            "Message Quality",
            options=[1, 2, 3, 4, 5],
            value=3,
            format_func=lambda x: {1: "1 — Poor", 2: "2 — Fair", 3: "3 — OK", 4: "4 — Good", 5: "5 — Excellent"}[x],
        )
        clinically_accurate = st.radio(
            "Clinically accurate content?", ["Yes", "No", "Partially"], horizontal=True
        )
    with fcol2:
        appropriate_tone = st.radio(
            "Appropriate tone?", ["Yes", "No"], horizontal=True
        )
        actionable = st.radio(
            "Actionable suggestion?", ["Yes", "No"], horizontal=True
        )

    feedback_text = st.text_area(
        "Additional feedback / notes",
        placeholder="e.g., The glucose mention was incorrect because...",
    )

    submit_btn = st.button("💾 Submit Feedback & Log to Excel", type="primary")

    if submit_btn:
        # Build the log row
        actions_str = "; ".join(
            [a["key"] for a in result["positive_actions_used"]]
        )
        opp_str = result["opportunity_used"]["key"]

        log_row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            tester_name,
            # Config
            a1c_group, user_focus, weight_goal,
            has_cgm, has_step_tracker, has_medications,
            glucose_scenario, step_scenario, activity_scenario,
            sleep_scenario, food_scenario, med_scenario,
            weight_scenario, mental_scenario, journey_scenario,
            exercise_video_scenario, exercise_program_scenario,
            # Model params
            temperature, max_tokens, max_actions, greeting_override,
            # Results
            user_ctx.patient_id, result["rating"], result["rating_description"],
            actions_str, opp_str,
            result["message"], result["character_count"], result["word_count"],
            result["success"],
            # Feedback
            quality_rating, clinically_accurate,
            appropriate_tone, actionable,
            feedback_text,
        ]

        _append_log_row(log_row)
        st.session_state.logged = True
        st.success("✅ Feedback logged successfully! You can download the log from the sidebar.")

    if st.session_state.logged:
        st.info("This test run has been logged. Generate a new insight to start a fresh test.")

else:
    # Landing state
    st.markdown("---")
    st.markdown(
        """
        ### How to use this tool
        
        1. **Configure** a synthetic patient profile using the dropdowns in the sidebar
        2. **Adjust** health metric scenarios to create the test case you want
        3. Click **🚀 Generate Insight** to run the pipeline
        4. **Review** the generated message, selected actions, and score breakdown
        5. **Submit feedback** — your rating and notes are saved to an Excel log
        6. **Download** the log anytime from the sidebar
        
        ---
        
        **Scenario dropdowns control the synthetic data:**
        
        | Dropdown | What it controls |
        |----------|------------------|
        | Glucose Performance | TIR %, high/low glucose percentages |
        | Daily Steps | Step count range |
        | Weekly Activity | Active minutes (daily + weekly) |
        | Sleep Quality | Sleep hours + sleep rating |
        | Food Logging | Meals logged, nutrient targets |
        | Medication Adherence | Adherence %, took all meds flag |
        | Weight Trend | Weight change, logging frequency |
        | Mental Wellbeing | Meditation, journaling, action plans |
        | Journey | Active journey, task completion |
        | Exercise Video | Video completion status |
        | Exercise Program | Program activity and progress |
        """
    )
