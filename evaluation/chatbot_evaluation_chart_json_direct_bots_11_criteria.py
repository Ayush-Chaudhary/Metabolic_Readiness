# Databricks notebook source
# MAGIC %md
# MAGIC # Comprehensive Chatbot Response Evaluation - Configurable Quality Criteria
# MAGIC ## Multi-Criteria Quality Assessment using Bot Chains Directly
# MAGIC
# MAGIC This notebook evaluates chatbot responses across **configurable quality criteria** by **calling bot chains directly** (no ChatAgent wrapper).
# MAGIC
# MAGIC **🎯 NEW: Fully Configurable Evaluation Metrics**
# MAGIC - Choose which criteria to evaluate (evaluate 1, 5, 10, or all criteria)
# MAGIC - Organize criteria into custom categories
# MAGIC - Two methods: simple configuration in Section 4, or programmatic in Section 11
# MAGIC
# MAGIC **Default Evaluation Criteria (11 total):**
# MAGIC
# MAGIC **ACCURACY (1 criterion):**
# MAGIC 1. data_context_accuracy - Numeric data from source context correctly transferred to response
# MAGIC
# MAGIC **COACHING & CONTENT (1 criterion):**
# MAGIC 2. coaching_appropriateness - Guidance within boundaries
# MAGIC
# MAGIC **TONE & STYLE (4 criteria):**
# MAGIC 3. tone_warmth_and_approachability - Friendly and supportive
# MAGIC 4. tone_professionalism - Credible and appropriate
# MAGIC 5. reading_level_appropriateness - 5th-6th grade level
# MAGIC 6. message_length_appropriateness - Under 350 words
# MAGIC
# MAGIC **LANGUAGE & TERMINOLOGY (4 criteria):**
# MAGIC 7. spelling_and_punctuation - Correct spelling and punctuation rules
# MAGIC 8. medical_terminology_correctness - Proper medical terminology
# MAGIC 9. inappropriate_language_avoidance - No stigmatizing language
# MAGIC 10. language_variety_and_flow - Varied and natural language
# MAGIC
# MAGIC **FORMAT & PRESENTATION (1 criterion):**
# MAGIC 11. response_structure_and_format - Clear organization
# MAGIC
# MAGIC **Key Features:**
# MAGIC - ✨ **NEW: Fully configurable criteria** - choose which metrics to evaluate
# MAGIC - Calls bot chains directly (no UnifiedChatAgentV3 wrapper)
# MAGIC - Supports optional `user_id`, `patient_name`, and `env` columns
# MAGIC - Routes queries to appropriate bots
# MAGIC - Chart rendering disabled for faster evaluation
# MAGIC - Generates detailed Excel reports
# MAGIC - Validation checks for configuration consistency

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. Configuration - UPDATE THESE PATHS

# COMMAND ----------

# MAGIC %load_ext autoreload
# MAGIC %autoreload 2
# MAGIC
# MAGIC %pip install -U --quiet \
# MAGIC     mlflow==3.1.4 \
# MAGIC     langchain>=0.3.0 \
# MAGIC     langchain-core>=0.3.0 \
# MAGIC     langchain-community \
# MAGIC     databricks-langchain>=0.6.0 \
# MAGIC     databricks-agents \
# MAGIC     databricks-sdk>=0.64.0 \
# MAGIC     unitycatalog-ai \
# MAGIC     databricks-sql-connector==4.0.5 \
# MAGIC     playwright==1.40.0 \
# MAGIC     openpyxl \
# MAGIC     Pillow \
# MAGIC     nest_asyncio \
# MAGIC     matplotlib
# MAGIC
# MAGIC dbutils.library.restartPython()

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1.1 Install System Dependencies for Playwright (Required)

# COMMAND ----------

# Install system dependencies required for Playwright browsers
# This needs to be run with sudo in Databricks
import subprocess
import sys

try:
    # Try to install Playwright dependencies
    result = subprocess.run(
        ["playwright", "install-deps"], 
        capture_output=True, 
        text=True,
        timeout=300
    )
    if result.returncode == 0:
        print("✅ Playwright system dependencies installed successfully")
    else:
        print("⚠️  Could not install with playwright install-deps")
        print("Attempting alternative installation...")
        
        # Alternative: Install specific packages
        packages = [
            "libatk1.0-0",
            "libatk-bridge2.0-0", 
            "libatspi2.0-0",
            "libxdamage1",
            "libxcomposite1",
            "libxrandr2",
            "libgbm1",
            "libpango-1.0-0",
            "libcairo2",
            "libasound2"
        ]
        
        # Note: This may require cluster admin permissions
        print("Installing system packages (may require admin permissions)...")
        for package in packages:
            subprocess.run(["apt-get", "install", "-y", package], capture_output=True)
        
        print("✅ System packages installed")
        
except subprocess.TimeoutExpired:
    print("⚠️  Installation timed out")
except Exception as e:
    print(f"⚠️  Error installing dependencies: {e}")
    print("\nIf you see permission errors, you have two options:")
    print("1. Contact your Databricks admin to install system dependencies")
    print("2. Use matplotlib-based chart generation instead (see alternative below)")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1.2 Install Playwright Browser (One-time setup)

# COMMAND ----------

# Install Playwright browsers (run this once)
# This installs Chromium for headless chart rendering
import subprocess
try:
    subprocess.run(["playwright", "install", "chromium"], check=True)
    print("✅ Playwright Chromium browser installed")
except Exception as e:
    print(f"⚠️  Playwright browser installation: {e}")
    print("Run: playwright install chromium")

# COMMAND ----------

# ========================================================================
# CRITICAL: UPDATE THESE PATHS BEFORE RUNNING
# ========================================================================

# Directory containing all config YAML files and bot modules
CONFIG_SOURCE_DIR = "/Workspace/Users/achaudhary@welldocinc.com/Welldoc_3_4/Chatbot_V3_Restructured/configs"
BOTS_MODULE_DIR = "/Workspace/Users/achaudhary@welldocinc.com/Welldoc_3_4/Chatbot_V3_Restructured"

# Path to evaluation guidelines YAML (11 criteria - chart criteria disabled)
GUIDELINES_PATH = "evaluation_guidelines.yaml"

# Judge model for evaluation (Claude Sonnet 4)
JUDGE_MODEL_NAME = "databricks-claude-sonnet-4"

# Default patient info (used if not provided in data)
DEFAULT_PATIENT_ID = "12345"
DEFAULT_PATIENT_NAME = "there"

# Chart rendering settings
CHART_IMAGE_DIR = "chart_images"  # Directory to store rendered chart PNGs
CHART_WIDTH = 800  # Width of rendered chart images in pixels
CHART_HEIGHT = 400  # Height of rendered chart images in pixels

print("✅ Configuration set")
print(f"Config Directory: {CONFIG_SOURCE_DIR}")
print(f"Bots Module Directory: {BOTS_MODULE_DIR}")
print(f"Guidelines: {GUIDELINES_PATH}")
print(f"Judge Model: {JUDGE_MODEL_NAME}")
print(f"Chart Image Directory: {CHART_IMAGE_DIR}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. Import Required Libraries

# COMMAND ----------

import mlflow
import pandas as pd
import numpy as np
import yaml
from typing import Dict, List, Any, Optional
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
import uuid
import openai
import os
import sys

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Mute verbose loggers
logging.getLogger("py4j").setLevel(logging.WARNING)
logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)

# Enable MLflow's autologging
mlflow.openai.autolog()

# Set up MLflow tracking
mlflow.set_tracking_uri("databricks")
mlflow.set_experiment("/Shared/llm-evaluation-chatbot-quality-11-criteria")

print("✅ Libraries imported successfully")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 3. Import Bot Modules Directly

# COMMAND ----------

# Add bots directory to path
if BOTS_MODULE_DIR not in sys.path:
    sys.path.insert(0, BOTS_MODULE_DIR)
    print(f"✅ Added {BOTS_MODULE_DIR} to Python path")

# Import bot classes and utilities from the new unified structure
# All bots are now in bots.py, config utilities in config_utils.py
from bots import (
    MedicationBot, DietBot, ActivityBot,
    LabsBGBot, LabsBPBot, LabsGlucoseBot, LabsHFBot, LabsWeightBot,
    SleepBot,
    EducationBot, FAQBot, RecipeBot, RouterBot
)
from config_utils import set_config_dir, load_merged_config
from mlflow.types.agent import ChatAgentMessage

print("✅ Successfully imported all bot classes from bots.py")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 4. Define Evaluation Criteria (Configurable)

# COMMAND ----------

# ========================================================================
# EVALUATION CRITERIA CONFIGURATION
# ========================================================================
# INSTRUCTIONS: Customize this section to select which criteria you want to evaluate.
# 
# Option 1: Use default criteria (11 criteria - chart evaluation disabled)
# Option 2: Specify custom criteria list and categories
# 
# To customize:
# 1. Modify EVALUATION_CRITERIA list to include only the criteria you want
# 2. Update CRITERIA_CATEGORIES to organize your criteria into logical groups
# 3. Ensure all criteria in EVALUATION_CRITERIA are defined in evaluation_guidelines.yaml
# ========================================================================

# DEFAULT CRITERIA SET (11 criteria - chart evaluation disabled)
DEFAULT_EVALUATION_CRITERIA = [
    # ACCURACY CRITERIA (1 - chart-related criteria disabled)
    "data_context_accuracy",
    
    # COACHING & CONTENT CRITERIA (1)
    "coaching_appropriateness",
    
    # TONE & STYLE CRITERIA (4)
    "tone_warmth_and_approachability",
    "tone_professionalism",
    "reading_level_appropriateness",
    "message_length_appropriateness",
    
    # LANGUAGE & TERMINOLOGY CRITERIA (4)
    "spelling_and_punctuation",
    "medical_terminology_correctness",
    "inappropriate_language_avoidance",
    "language_variety_and_flow",
    
    # FORMAT & PRESENTATION CRITERIA (1)
    "response_structure_and_format",
]

DEFAULT_CRITERIA_CATEGORIES = {
    "Accuracy": [
        "data_context_accuracy"
    ],
    "Coaching & Content": [
        "coaching_appropriateness",
    ],
    "Tone & Style": [
        "tone_warmth_and_approachability",
        "tone_professionalism",
        "reading_level_appropriateness",
        "message_length_appropriateness"
    ],
    "Language & Terminology": [
        "spelling_and_punctuation",
        "medical_terminology_correctness",
        "inappropriate_language_avoidance",
        "language_variety_and_flow"
    ],
    "Format & Presentation": [
        "response_structure_and_format",
    ]
}

# ========================================================================
# CUSTOM CRITERIA CONFIGURATION (OPTIONAL)
# ========================================================================
# Uncomment and modify this section to use custom evaluation criteria
# ========================================================================

# Example: Evaluate only accuracy and tone criteria
CUSTOM_EVALUATION_CRITERIA = [
    # "data_context_accuracy",
    # COACHING & CONTENT CRITERIA (1)
    "coaching_appropriateness",
    
    # TONE & STYLE CRITERIA (4)
    # "tone_warmth_and_approachability",
    # "tone_professionalism",
    "reading_level_appropriateness",
    # "message_length_appropriateness",
    
    # LANGUAGE & TERMINOLOGY CRITERIA (4)
    "spelling_and_punctuation",
    # "medical_terminology_correctness",
    # "inappropriate_language_avoidance",
    # "language_variety_and_flow",
    
    # FORMAT & PRESENTATION CRITERIA (1)
    # "response_structure_and_format",
]

CUSTOM_CRITERIA_CATEGORIES = {
    # "Accuracy": ["data_context_accuracy"],
    "Coaching & Content": [
        "coaching_appropriateness",
    ],
    "Tone & Style": [
        # "tone_warmth_and_approachability",
        # "tone_professionalism",
        "reading_level_appropriateness",
        # "message_length_appropriateness"
    ],
    "Language & Terminology": [
        "spelling_and_punctuation",
        # "medical_terminology_correctness",
        # "inappropriate_language_avoidance",
        # "language_variety_and_flow"
    ],
    # "Format & Presentation": [
    #     "response_structure_and_format",
    # ]
}

# ========================================================================
# SELECT WHICH CRITERIA TO USE
# ========================================================================
# Set USE_CUSTOM_CRITERIA = True to use custom criteria defined above
# Set USE_CUSTOM_CRITERIA = False to use default criteria
USE_CUSTOM_CRITERIA = True

if USE_CUSTOM_CRITERIA and 'CUSTOM_EVALUATION_CRITERIA' in locals():
    EVALUATION_CRITERIA = CUSTOM_EVALUATION_CRITERIA
    CRITERIA_CATEGORIES = CUSTOM_CRITERIA_CATEGORIES
    print("✅ Using CUSTOM evaluation criteria")
else:
    EVALUATION_CRITERIA = DEFAULT_EVALUATION_CRITERIA
    CRITERIA_CATEGORIES = DEFAULT_CRITERIA_CATEGORIES
    print("✅ Using DEFAULT evaluation criteria")

# Validate criteria configuration
print(f"\n📋 Evaluating {len(EVALUATION_CRITERIA)} criteria across {len(CRITERIA_CATEGORIES)} categories:")
for category, criteria in CRITERIA_CATEGORIES.items():
    print(f"\n  {category} ({len(criteria)} criteria):")
    for criterion in criteria:
        print(f"    - {criterion.replace('_', ' ').title()}")

# Validation check
all_category_criteria = [c for criteria_list in CRITERIA_CATEGORIES.values() for c in criteria_list]
missing_from_categories = set(EVALUATION_CRITERIA) - set(all_category_criteria)
extra_in_categories = set(all_category_criteria) - set(EVALUATION_CRITERIA)

if missing_from_categories:
    print(f"\n⚠️  WARNING: These criteria are in EVALUATION_CRITERIA but not in CRITERIA_CATEGORIES:")
    for criterion in missing_from_categories:
        print(f"    - {criterion}")

if extra_in_categories:
    print(f"\n⚠️  WARNING: These criteria are in CRITERIA_CATEGORIES but not in EVALUATION_CRITERIA:")
    for criterion in extra_in_categories:
        print(f"    - {criterion}")

if not missing_from_categories and not extra_in_categories:
    print("\n✅ Criteria configuration is valid")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 5. Configuration Classes

# COMMAND ----------

@dataclass
class Criterion:
    name: str
    is_labeled: bool
    guideline: str

@dataclass
class EvaluationConfig:
    """Configuration for evaluation pipeline with customizable criteria"""
    config_source_dir: str
    judge_model_name: str = "databricks-claude-sonnet-4"
    temperature: float = 0.0
    max_tokens: int = 4000  # For detailed criteria analysis
    guidelines_path: str = "evaluation_guidelines.yaml"
    relevant_criteria: List[str] = field(default_factory=lambda: EVALUATION_CRITERIA)
    criteria_categories: Dict[str, List[str]] = field(default_factory=lambda: CRITERIA_CATEGORIES)
    unlabeled_criteria: List[str] = field(default_factory=list)
    labeled_criteria: List[str] = field(default_factory=list)
    guideline_map: Dict[str, str] = field(default_factory=dict)

    def __post_init__(self):
        """Initialize configuration after creation."""
        self.load_guidelines()
        self._validate_criteria()

    def load_guidelines(self):
        """Load evaluation criteria and guidelines from YAML file."""
        try:
            with open(self.guidelines_path, 'r') as f:
                data = yaml.safe_load(f)
            
            for item in data.get("evaluation_criteria", []):
                criterion = Criterion(**item)
                if criterion.name in self.relevant_criteria:
                    self.guideline_map[criterion.name] = criterion.guideline
                    if criterion.is_labeled:
                        self.labeled_criteria.append(criterion.name)
                    else:
                        self.unlabeled_criteria.append(criterion.name)
            
            logger.info(f"✅ Loaded {len(self.guideline_map)} guidelines")
            logger.info(f"   - Labeled criteria: {len(self.labeled_criteria)}")
            logger.info(f"   - Unlabeled criteria: {len(self.unlabeled_criteria)}")
            
            # Check for missing guidelines
            missing_guidelines = set(self.relevant_criteria) - set(self.guideline_map.keys())
            if missing_guidelines:
                logger.warning(f"⚠️  WARNING: Missing guidelines for criteria: {missing_guidelines}")
            
        except FileNotFoundError:
            logger.error(f"❌ Guidelines file not found: {self.guidelines_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error loading guidelines: {e}")
            raise
    
    def _validate_criteria(self):
        """Validate that criteria configuration is consistent."""
        all_category_criteria = [c for criteria_list in self.criteria_categories.values() for c in criteria_list]
        
        missing_from_categories = set(self.relevant_criteria) - set(all_category_criteria)
        extra_in_categories = set(all_category_criteria) - set(self.relevant_criteria)
        
        if missing_from_categories:
            logger.warning(f"⚠️  Criteria in relevant_criteria but not in categories: {missing_from_categories}")
        
        if extra_in_categories:
            logger.warning(f"⚠️  Criteria in categories but not in relevant_criteria: {extra_in_categories}")

def create_custom_evaluation_config(
    config_source_dir: str,
    evaluation_criteria: List[str],
    criteria_categories: Dict[str, List[str]],
    judge_model_name: str = "databricks-claude-sonnet-4",
    guidelines_path: str = "evaluation_guidelines.yaml"
) -> EvaluationConfig:
    """
    Helper function to create an EvaluationConfig with custom criteria.
    
    Args:
        config_source_dir: Directory containing bot config YAML files
        evaluation_criteria: List of criterion names to evaluate
        criteria_categories: Dict mapping category names to lists of criteria
        judge_model_name: Name of the judge model (default: databricks-claude-sonnet-4)
        guidelines_path: Path to evaluation guidelines YAML file
    
    Returns:
        EvaluationConfig object with custom criteria
    
    Example:
        config = create_custom_evaluation_config(
            config_source_dir="/path/to/configs",
            evaluation_criteria=[
                "data_context_accuracy",
                "tone_warmth_and_approachability",
                "medical_terminology_correctness"
            ],
            criteria_categories={
                "Accuracy": ["data_context_accuracy"],
                "Tone": ["tone_warmth_and_approachability"],
                "Language": ["medical_terminology_correctness"]
            }
        )
    """
    return EvaluationConfig(
        config_source_dir=config_source_dir,
        judge_model_name=judge_model_name,
        guidelines_path=guidelines_path,
        relevant_criteria=evaluation_criteria,
        criteria_categories=criteria_categories
    )

print("✅ Configuration classes defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 6. Chart Renderer Class (PNG Generation)

# COMMAND ----------

from playwright.async_api import async_playwright
from pathlib import Path
import base64
import asyncio

class ChartRenderer:
    """Renders Highcharts JSON to PNG images using Playwright (async)."""
    
    def __init__(self, output_dir: str = CHART_IMAGE_DIR, width: int = CHART_WIDTH, height: int = CHART_HEIGHT):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.width = width
        self.height = height
        
    async def _render_chart_async(self, chart_json: Dict[str, Any], filename: str) -> str:
        """
        Async method to render Highcharts JSON to PNG using Playwright.
        
        Args:
            chart_json: Highcharts configuration JSON
            filename: Output filename (e.g., "chart_1.png")
            
        Returns:
            Path to saved PNG file
        """
        output_path = self.output_dir / filename
        
        # Create HTML with Highcharts
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <script src="https://code.highcharts.com/highcharts.js"></script>
            <script src="https://code.highcharts.com/modules/exporting.js"></script>
            <script src="https://code.highcharts.com/modules/export-data.js"></script>
        </head>
        <body style="margin: 0; padding: 0;">
            <div id="container" style="width: {self.width}px; height: {self.height}px;"></div>
            <script>
                const chartConfig = {json.dumps(chart_json)};
                chartConfig.chart = chartConfig.chart || {{}};
                chartConfig.chart.renderTo = 'container';
                chartConfig.chart.width = {self.width};
                chartConfig.chart.height = {self.height};
                Highcharts.chart('container', chartConfig);
            </script>
        </body>
        </html>
        """
        
        try:
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                page = await browser.new_page(viewport={'width': self.width, 'height': self.height})
                await page.set_content(html_content)
                await page.wait_for_timeout(2000)  # Wait for chart to render
                await page.screenshot(path=str(output_path))
                await browser.close()
            
            logger.info(f"✅ Chart rendered to: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"❌ Error rendering chart: {e}")
            return None
    
    def render_chart_to_png(self, chart_json: Dict[str, Any], filename: str) -> str:
        """
        Render Highcharts JSON to PNG using Playwright (sync wrapper for async).
        
        Args:
            chart_json: Highcharts configuration JSON
            filename: Output filename (e.g., "chart_1.png")
            
        Returns:
            Path to saved PNG file
        """
        # Check if we're already in an event loop
        try:
            loop = asyncio.get_running_loop()
            # We're in an event loop, use nest_asyncio or run in thread
            import nest_asyncio
            nest_asyncio.apply()
            return asyncio.run(self._render_chart_async(chart_json, filename))
        except RuntimeError:
            # No event loop, safe to use asyncio.run
            return asyncio.run(self._render_chart_async(chart_json, filename))

print("✅ ChartRenderer class defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 6.1 Alternative Chart Renderer (Matplotlib - No System Dependencies)

# COMMAND ----------

import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime as dt

class MatplotlibChartRenderer:
    """Alternative chart renderer using matplotlib (no system dependencies needed)."""
    
    def __init__(self, output_dir: str = CHART_IMAGE_DIR, width: int = 10, height: int = 6):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.width = width
        self.height = height
        
    def render_chart_to_png(self, chart_config: Dict[str, Any], filename: str) -> str:
        """
        Render chart JSON to PNG using matplotlib.
        
        Args:
            chart_config: Highcharts configuration JSON
            filename: Output filename
            
        Returns:
            Path to saved PNG file
        """
        output_path = self.output_dir / filename
        
        try:
            # Extract data from Highcharts config
            series_data = chart_config.get('series', [])
            if not series_data:
                logger.warning("No series data found in chart config")
                return None
            
            chart_type = chart_config.get('chart', {}).get('type', 'line')
            title = chart_config.get('title', {}).get('text', 'Chart')
            x_axis_label = chart_config.get('xAxis', {}).get('title', {}).get('text', '')
            y_axis_label = chart_config.get('yAxis', {}).get('title', {}).get('text', '')
            
            # Create figure
            fig, ax = plt.subplots(figsize=(self.width, self.height))
            
            # Plot each series
            for series in series_data:
                data = series.get('data', [])
                if not data:
                    continue
                
                series_name = series.get('name', 'Data')
                
                # Extract x and y values
                if isinstance(data[0], list):
                    # Data is [[x1, y1], [x2, y2], ...]
                    x_values = [point[0] for point in data]
                    y_values = [point[1] for point in data]
                    
                    # Convert timestamps to dates if needed
                    if all(isinstance(x, (int, float)) and x > 1000000000000 for x in x_values):
                        # Unix timestamps in milliseconds
                        x_values = [dt.fromtimestamp(x / 1000) for x in x_values]
                        ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %d'))
                        plt.xticks(rotation=45)
                else:
                    # Data is [y1, y2, y3, ...]
                    y_values = data
                    x_values = list(range(len(y_values)))
                
                # Plot based on chart type
                if chart_type in ['line', 'spline']:
                    ax.plot(x_values, y_values, marker='o', label=series_name, linewidth=2)
                elif chart_type in ['column', 'bar']:
                    ax.bar(x_values, y_values, label=series_name)
                else:
                    ax.plot(x_values, y_values, label=series_name)
            
            # Set labels and title
            ax.set_title(title, fontsize=14, fontweight='bold')
            if x_axis_label:
                ax.set_xlabel(x_axis_label, fontsize=11)
            if y_axis_label:
                ax.set_ylabel(y_axis_label, fontsize=11)
            
            # Add legend if multiple series
            if len(series_data) > 1:
                ax.legend()
            
            # Grid
            ax.grid(True, alpha=0.3)
            
            # Tight layout
            plt.tight_layout()
            
            # Save
            plt.savefig(output_path, dpi=100, bbox_inches='tight')
            plt.close()
            
            logger.info(f"✅ Chart rendered with matplotlib to: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"❌ Error rendering chart with matplotlib: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None

print("✅ MatplotlibChartRenderer class defined (fallback option)")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 7. Direct Bot Manager

# COMMAND ----------

class DirectBotManager:
    """Manages bot initialization and routing for direct bot invocation."""
    
    def __init__(self, config_dir: str):
        """Initialize all bots."""
        self.config_dir = config_dir
        set_config_dir(config_dir)
        
        # Load config files using mlflow.models.ModelConfig
        logger.info("🤖 Loading configuration files...")
        
        personalized_config_path = os.path.join(config_dir, "personalized_bot_config.yaml")
        router_config_path = os.path.join(config_dir, "router_config_updated.yaml")
        education_config_path = os.path.join(config_dir, "education_rag_chain_config.yaml")
        faq_config_path = os.path.join(config_dir, "faq_rag_chain_config.yaml")
        recipe_config_path = os.path.join(config_dir, "recipe_rag_chain_config.yaml")
        
        personalized_config = mlflow.models.ModelConfig(development_config=personalized_config_path)
        router_config = mlflow.models.ModelConfig(development_config=router_config_path)
        education_config = mlflow.models.ModelConfig(development_config=education_config_path)
        faq_config = mlflow.models.ModelConfig(development_config=faq_config_path)
        recipe_config = mlflow.models.ModelConfig(development_config=recipe_config_path)
        
        # Initialize all bots with proper config objects
        logger.info("🤖 Initializing bots...")
        
        self.bots = {
            "medication": MedicationBot(personalized_config),
            "diet": DietBot(personalized_config),
            "activity": ActivityBot(personalized_config),
            "labs_bg": LabsBGBot(personalized_config),
            "labs_bp": LabsBPBot(personalized_config),
            "labs_glucose": LabsGlucoseBot(personalized_config),
            "labs_hf": LabsHFBot(personalized_config),
            "labs_weight": LabsWeightBot(personalized_config),
            "sleep": SleepBot(personalized_config),
            "education": EducationBot(education_config),
            "product_faq": FAQBot(faq_config),
            "recipe": RecipeBot(recipe_config)
        }
        
        self.router_bot = RouterBot(router_config)
        logger.info(f"✅ Initialized {len(self.bots)} bots + router")
        
    def route_query(self, query: str) -> str:
        """Route query to appropriate bot using router_bot."""
        messages = [ChatAgentMessage(content=query, role="user")]
        try:
            # router_decision returns a string directly (the route name), not a dict
            route = self.router_bot.router_decision(messages)
            logger.info(f"🧭 Routed to: {route}")
            return route
        except Exception as e:
            logger.error(f"❌ Routing error: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return "product_faq"  # Default fallback
            
    def invoke_bot(self, route: str, query: str, user_id: str = None, patient_name: str = None, env: str = "uat") -> Dict[str, Any]:
        """Invoke bot chain directly."""
        
        # Handle special 'clarify' route - not a bot, returns predefined response
        if route == "clarify":
            logger.info("🔍 Clarify route - returning clarification request")
            return {
                "response": "I need a bit more information to help you properly. Could you please clarify:\n"
                           "- Are you asking about general health information?\n"
                           "- Or about your personal health data (like 'my glucose', 'my medications')?\n"
                           "- Or about how to use the app?\n\n"
                           "Please rephrase your question with more details.",
                "chart_config": {},
                "show_chart": False,
                "data_context": {},
                "classification": "clarify",
                "confidence": 1.0
            }
        
        bot = self.bots.get(route)
        if not bot:
            logger.error(f"❌ Bot not found for route: {route}")
            return {
                "response": "I apologize, but I'm unable to process that query right now.",
                "chart_config": {},
                "show_chart": False,
                "data_context": {}
            }
        
        # Prepare messages
        messages = [ChatAgentMessage(content=query, role="user")]
        
        # Prepare custom inputs - must include env for data fetching
        custom_inputs = {
            "env": env  # Required for connecting to correct environment
        }
        if user_id:
            custom_inputs["patient_id"] = user_id
        if patient_name:
            custom_inputs["patient_name"] = patient_name
            
        try:
            # Invoke bot chain
            chain = bot.get_chain()
            result = chain.invoke({
                "messages": messages,
                "custom_inputs": custom_inputs
            })
            
            return {
                "response": result.get("response", ""),
                "chart_config": result.get("chart_config", {}),
                "show_chart": result.get("show_chart", False),
                "data_context": result.get("data_context", {}),
                "classification": result.get("classification"),
                "confidence": result.get("confidence"),
                "reasoning": result.get("reasoning")
            }
            
        except Exception as e:
            logger.error(f"❌ Bot invocation error: {e}")
            return {
                "response": "I apologize, but I encountered an error processing your query.",
                "chart_config": {},
                "show_chart": False,
                "data_context": {}
            }

print("✅ DirectBotManager class defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 8. Response Evaluator (Configurable Criteria)

# COMMAND ----------

class ResponseEvaluator:
    """Evaluates chatbot responses using LLM judge across configurable quality criteria."""
    
    def __init__(self, config: EvaluationConfig):
        self.config = config
        # Use Databricks WorkspaceClient for OpenAI-compatible client
        from databricks.sdk import WorkspaceClient
        w = WorkspaceClient()
        self.client = w.serving_endpoints.get_open_ai_client()
        logger.info("✅ Using Databricks WorkspaceClient for judge model")
        
    def evaluate_single_response(
        self, 
        query: str, 
        response: str, 
        chart_config: Dict[str, Any] = None,
        show_chart: bool = False,
        data_context: Dict[str, Any] = None
    ) -> Dict[str, Dict[str, str]]:
        """
        Evaluate a single response across all 11 criteria.
        
        Returns:
            Dict mapping criterion name to {"assessment": "yes"/"no", "rationale": "..."}
        """
        evaluations = {}
        
        # Format chart info for evaluation
        chart_info = ""
        if show_chart and chart_config:
            chart_info = f"\n\n**CHART JSON:**\n```json\n{json.dumps(chart_config, indent=2)}\n```"
        elif not show_chart:
            chart_info = "\n\n**NOTE:** This response has no chart (text-only response)."
        
        # Format data context for evaluation
        data_context_info = ""
        if data_context:
            data_context_info = f"\n\n**SOURCE DATA CONTEXT:**\n```json\n{json.dumps(data_context, indent=2, default=str)}\n```"
        else:
            data_context_info = "\n\n**NOTE:** No data context provided for this response."
        
        full_response_with_chart = response + chart_info
        
        for criterion in self.config.unlabeled_criteria:
            guideline = self.config.guideline_map.get(criterion, "")
            
            assessment, rationale = self._evaluate_criterion(
                query, 
                full_response_with_chart,
                criterion,
                guideline,
                data_context_info
            )
            # assessment, rationale = "hey", "there"
            
            evaluations[criterion] = {
                "assessment": assessment,
                "rationale": rationale
            }
            
        return evaluations
    
    def _evaluate_criterion(
        self, 
        query: str, 
        response_with_chart: str,
        criterion: str,
        guideline: str,
        data_context_info: str = ""
    ) -> tuple[str, str]:
        """Evaluate a single criterion using LLM judge."""
        
        # Include data context in prompt for data_context_accuracy criterion
        # and optionally for other accuracy-related criteria
        context_section = ""
        if criterion in ["data_context_accuracy", "numerical_accuracy", "data_interpretation_accuracy"]:
            context_section = data_context_info
        
        prompt = f"""You are an expert evaluator assessing chatbot responses for healthcare applications.

**User Query:**
{query}

**Chatbot Response (with chart JSON if applicable):**
{response_with_chart}
{context_section}

**Evaluation Criterion:** {criterion.replace('_', ' ').title()}

**Evaluation Guidelines:**
{guideline}

**Your Task:**
Carefully evaluate whether the response meets the criterion based on the guidelines provided.
{f"Pay special attention to verifying that numeric values in the response match the source data context exactly." if criterion == "data_context_accuracy" else ""}

**Response Format:**
Provide your evaluation in exactly this format:

ASSESSMENT: [yes or no]
RATIONALE: [1-2 sentences explaining your assessment]

Be objective and thorough. Base your assessment strictly on the guidelines provided."""

        try:
            completion = self.client.chat.completions.create(
                model=self.config.judge_model_name,
                messages=[{"role": "user", "content": prompt}],
                temperature=self.config.temperature,
                max_tokens=self.config.max_tokens
            )
            
            result = completion.choices[0].message.content.strip()
            
            # Parse response
            assessment_line = [line for line in result.split('\n') if line.startswith('ASSESSMENT:')]
            rationale_line = [line for line in result.split('\n') if line.startswith('RATIONALE:')]
            
            assessment = "no"  # Default
            if assessment_line:
                assessment_text = assessment_line[0].replace('ASSESSMENT:', '').strip().lower()
                assessment = "yes" if "yes" in assessment_text else "no"
            
            rationale = "No rationale provided"
            if rationale_line:
                rationale = rationale_line[0].replace('RATIONALE:', '').strip()
            
            return assessment, rationale
            
        except Exception as e:
            logger.error(f"❌ Error evaluating {criterion}: {e}")
            return "no", f"Evaluation error: {str(e)}"

print("✅ ResponseEvaluator class defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 9. Evaluation Pipeline

# COMMAND ----------

class EvaluationPipeline:
    """Orchestrates the complete evaluation workflow."""
    
    def __init__(self, config: EvaluationConfig):
        self.config = config
        self.bot_manager = DirectBotManager(config.config_source_dir)
        self.evaluator = ResponseEvaluator(config)
        
        # Use matplotlib as default (no system dependencies required)
        logger.info("📊 Using matplotlib for chart rendering (default)")
        self.chart_renderer = MatplotlibChartRenderer()
        
        # Uncomment below to use Playwright instead (requires system dependencies)
        # try:
        #     self.chart_renderer = ChartRenderer()
        #     logger.info("✅ Using Playwright for chart rendering")
        # except Exception as e:
        #     logger.warning(f"⚠️  Playwright not available: {e}")
        #     logger.info("📊 Falling back to matplotlib")
        #     self.chart_renderer = MatplotlibChartRenderer()
        
    def evaluate_dataset(
        self, 
        data: pd.DataFrame,
        user_id_col: str = "user_id",
        patient_name_col: str = "patient_name",
        query_col: str = "query",
        env_col: str = "env"
    ) -> List[Dict[str, Any]]:
        """
        Evaluate a dataset of queries.
        
        Args:
            data: DataFrame with queries
            user_id_col: Column name for user IDs (optional)
            patient_name_col: Column name for patient names (optional)
            query_col: Column name for queries
            env_col: Column name for environment (optional, defaults to "uat")
            
        Returns:
            List of evaluation results
        """
        results = []
        total = len(data)
        
        logger.info(f"📊 Starting evaluation of {total} queries...")
        
        for idx, row in data.iterrows():
            logger.info(f"\n{'='*60}")
            logger.info(f"🔍 Query {idx+1}/{total}")
            logger.info(f"{'='*60}")
            
            query = row[query_col]
            user_id = row.get(user_id_col, DEFAULT_PATIENT_ID)
            patient_name = "there"#row.get(patient_name_col, DEFAULT_PATIENT_NAME)
            env = row.get(env_col, "uat")  # Default to "uat" if not specified
            
            logger.info(f"Query: {query}")
            logger.info(f"User ID: {user_id}")
            logger.info(f"Patient Name: {patient_name}")
            logger.info(f"Environment: {env}")
            
            # Route and invoke bot
            route = self.bot_manager.route_query(query)
            bot_response = self.bot_manager.invoke_bot(route, query, user_id, patient_name, env)
            
            response_text = bot_response["response"]
            chart_config = bot_response["chart_config"]
            data_context = bot_response["data_context"]
            show_chart = bot_response["show_chart"]
            classification = bot_response.get("classification")
            confidence = bot_response.get("confidence")
            reasoning = bot_response.get("reasoning")
            
            logger.info(f"Response length: {len(response_text)} chars")
            logger.info(f"Has chart: {show_chart}")
            if classification:
                logger.info(f"Classification: {classification} (confidence: {confidence})")
            
            # Chart rendering disabled - not evaluating chart criteria
            chart_image_path = None
            # if show_chart and chart_config:
            #     chart_filename = f"chart_{idx+1}_{uuid.uuid4().hex[:8]}.png"
            #     try:
            #         chart_image_path = self.chart_renderer.render_chart_to_png(
            #             chart_config, 
            #             chart_filename
            #         )
            #     except Exception as chart_error:
            #         logger.warning(f"⚠️  Could not render chart to PNG: {chart_error}")
            #         logger.warning("Continuing evaluation without chart image...")
            #         chart_image_path = None
            
            # Evaluate response
            logger.info("⚖️  Evaluating response...")
            evaluations = self.evaluator.evaluate_single_response(
                query, 
                response_text,
                chart_config,
                show_chart,
                data_context
            )
            
            # Count passes
            passes = sum(1 for e in evaluations.values() if e["assessment"] == "yes")
            pass_rate = passes / len(evaluations) if evaluations else 0
            
            logger.info(f"✅ Evaluation complete: {passes}/{len(evaluations)} criteria passed ({pass_rate:.1%})")
            
            # Store result
            results.append({
                "query": query,
                "user_id": user_id,
                "patient_name": patient_name,
                "env": env,
                "route": route,
                "classification": classification,
                "confidence": confidence,
                "reasoning": reasoning,
                "response": response_text,
                "data_context": data_context,
                "chart_config": chart_config,
                "show_chart": show_chart,
                "chart_image_path": chart_image_path,
                "evaluations": evaluations,
                "pass_rate": pass_rate
            })
        
        logger.info(f"\n{'='*60}")
        logger.info(f"✅ Evaluation complete: {total} queries processed")
        logger.info(f"{'='*60}")
        
        return results

print("✅ EvaluationPipeline class defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 10. Excel Report Generator (with embedded PNGs)

# COMMAND ----------

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

class ExcelReportGenerator:
    """Generates Excel reports with embedded chart images."""
    
    def __init__(self, criteria_categories: Dict[str, List[str]], evaluation_criteria: List[str]):
        self.criteria_categories = criteria_categories
        self.evaluation_criteria = evaluation_criteria
        
    def generate_report(self, results: List[Dict[str, Any]], output_filename: str):
        """Generate comprehensive Excel report with embedded chart images."""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluation Results"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Define columns
        base_columns = [
            "Query",
            "User ID",
            "Patient Name",
            "Env",
            "Route",
            "Classification",
            "Confidence",
            "Data Context",
            "Response",
            "Chart Image",  # For embedded PNG
            "Chart JSON",
            "Show Chart"
        ]
        
        # Add assessment and rationale columns for each criterion
        criterion_columns = []
        for criterion in self.evaluation_criteria:
            criterion_name = criterion.replace('_', ' ').title()
            criterion_columns.extend([
                f"{criterion_name} - Assessment",
                f"{criterion_name} - Rationale"
            ])
        
        all_columns = base_columns + criterion_columns
        
        # Write headers
        for col_idx, header in enumerate(all_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data
        current_row = 2
        for result in results:
            # Base data
            ws.cell(row=current_row, column=1, value=result["query"])
            ws.cell(row=current_row, column=2, value=result.get("user_id", ""))
            ws.cell(row=current_row, column=3, value=result.get("patient_name", ""))
            ws.cell(row=current_row, column=4, value=result.get("env", "uat"))
            ws.cell(row=current_row, column=5, value=result.get("route", ""))
            ws.cell(row=current_row, column=6, value=result.get("classification", ""))
            ws.cell(row=current_row, column=7, value=result.get("confidence", ""))
            # Convert data_context dict to string for Excel
            data_context_str = json.dumps(result.get("data_context", {}), indent=2, default=str) if result.get("data_context") else ""
            ws.cell(row=current_row, column=8, value=data_context_str)
            ws.cell(row=current_row, column=9, value=result["response"])
            
            # # Chart image (column 10) - embedded PNG
            # if result.get("chart_image_path") and os.path.exists(result["chart_image_path"]):
            #     try:
            #         img = XLImage(result["chart_image_path"])
            #         # Scale image to fit in cell (adjust as needed)
            #         img.width = 400
            #         img.height = 200
                    
            #         # Anchor image to cell
            #         cell_ref = f"J{current_row}"
            #         ws.add_image(img, cell_ref)
                    
            #         # Set row height to accommodate image
            #         ws.row_dimensions[current_row].height = 150
            #     except Exception as e:
            #         logger.error(f"❌ Error embedding image: {e}")
            #         ws.cell(row=current_row, column=10, value="[Image Error]")
            # else:
            #     ws.cell(row=current_row, column=10, value="[No Chart]")
            
            # Chart JSON (column 11)
            chart_json_str = json.dumps(result.get("chart_config", {}), indent=2) if result.get("chart_config") else ""
            ws.cell(row=current_row, column=11, value=chart_json_str)
            
            # Show Chart flag (column 12)
            ws.cell(row=current_row, column=12, value=result.get("show_chart", False))
            
            # Evaluation results
            col_idx = 13
            for criterion in self.evaluation_criteria:
                eval_data = result["evaluations"].get(criterion, {})
                assessment = eval_data.get("assessment", "no")
                rationale = eval_data.get("rationale", "")
                
                # Assessment cell
                assessment_cell = ws.cell(row=current_row, column=col_idx, value=assessment)
                if assessment == "yes":
                    assessment_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    assessment_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Rationale cell
                ws.cell(row=current_row, column=col_idx + 1, value=rationale)
                
                col_idx += 2
            
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40  # Query
        ws.column_dimensions['B'].width = 15  # User ID
        ws.column_dimensions['C'].width = 15  # Patient Name
        ws.column_dimensions['D'].width = 10  # Env
        ws.column_dimensions['E'].width = 15  # Route
        ws.column_dimensions['F'].width = 15  # Classification
        ws.column_dimensions['G'].width = 12  # Confidence
        ws.column_dimensions['H'].width = 50  # Data Context
        ws.column_dimensions['I'].width = 60  # Response
        ws.column_dimensions['J'].width = 55  # Chart Image
        ws.column_dimensions['K'].width = 40  # Chart JSON
        ws.column_dimensions['L'].width = 12  # Show Chart
        
        # Criterion columns
        for col_idx in range(13, len(all_columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30
        
        # Freeze panes
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(output_filename)
        logger.info(f"✅ Excel report saved: {output_filename}")

print("✅ ExcelReportGenerator class defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 11. Summary Generator (Configurable Criteria)

# COMMAND ----------

class SummaryGenerator:
    """Generates summary statistics and reports."""
    
    def __init__(self, criteria_categories: Dict[str, List[str]], evaluation_criteria: List[str]):
        self.criteria_categories = criteria_categories
        self.evaluation_criteria = evaluation_criteria
        
    def generate_summary(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate comprehensive summary statistics."""
        
        # Overall statistics
        total_queries = len(results)
        queries_with_charts = sum(1 for r in results if r.get("show_chart", False))
        queries_text_only = total_queries - queries_with_charts
        
        # Calculate per-criterion performance
        criterion_stats = {}
        for criterion in self.evaluation_criteria:
            scores = []
            for result in results:
                eval_data = result["evaluations"].get(criterion, {})
                score = 1 if eval_data.get("assessment") == "yes" else 0
                scores.append(score)
            
            if scores:
                criterion_stats[criterion] = {
                    "pass_rate": np.mean(scores),
                    "passes": sum(scores),
                    "total": len(scores)
                }
        
        # Calculate per-category performance
        category_stats = {}
        for category, criteria_list in self.criteria_categories.items():
            all_scores = []
            for criterion in criteria_list:
                if criterion in criterion_stats:
                    all_scores.extend([criterion_stats[criterion]["passes"], 
                                      criterion_stats[criterion]["total"] - criterion_stats[criterion]["passes"]])
            
            if all_scores:
                passes = sum(1 for i, score in enumerate(all_scores) if i % 2 == 0 and score > 0)
                total = len(all_scores) // 2
                category_stats[category] = {
                    "pass_rate": passes / total if total > 0 else 0,
                    "criteria_count": len(criteria_list)
                }
        
        # Overall pass rate
        all_scores = [score for stats in criterion_stats.values() 
                     for _ in range(stats["passes"])] + \
                    [0 for stats in criterion_stats.values() 
                     for _ in range(stats["total"] - stats["passes"])]
        overall_pass_rate = np.mean(all_scores) if all_scores else 0
        
        return {
            "total_queries": total_queries,
            "queries_with_charts": queries_with_charts,
            "queries_text_only": queries_text_only,
            "overall_pass_rate": overall_pass_rate,
            "criterion_stats": criterion_stats,
            "category_stats": category_stats,
            "timestamp": datetime.now().isoformat()
        }
    
    def print_summary(self, summary: Dict[str, Any]):
        """Print formatted summary to console."""
        
        num_criteria = len(self.evaluation_criteria)
        print("\n" + "="*80)
        print(f"📊 EVALUATION SUMMARY - {num_criteria} EVALUATION CRITERIA")
        print("="*80)
        
        print(f"\n📈 Overall Statistics:")
        print(f"   Total Queries: {summary['total_queries']}")
        print(f"   Queries with Charts: {summary['queries_with_charts']} "
              f"({summary['queries_with_charts']/summary['total_queries']:.1%})")
        print(f"   Text-Only Queries: {summary['queries_text_only']} "
              f"({summary['queries_text_only']/summary['total_queries']:.1%})")
        print(f"   Overall Pass Rate: {summary['overall_pass_rate']:.1%}")
        
        print(f"\n📋 Performance by Category:")
        print("-" * 80)
        for category, stats in summary["category_stats"].items():
            print(f"\n   {category} ({stats['criteria_count']} criteria):")
            print(f"      Pass Rate: {stats['pass_rate']:.1%}")
        
        print(f"\n📋 Performance by Criterion:")
        print("-" * 80)
        print(f"{'Criterion':<45} {'Pass Rate':<15} {'Passes/Total'}")
        print("-" * 80)
        
        for category, criteria_list in self.criteria_categories.items():
            print(f"\n{category}:")
            for criterion in criteria_list:
                if criterion in summary["criterion_stats"]:
                    stats = summary["criterion_stats"][criterion]
                    criterion_name = criterion.replace('_', ' ').title()
                    print(f"   {criterion_name:<42} {stats['pass_rate']:>6.1%}{' ':<8} "
                          f"{stats['passes']}/{stats['total']}")
        
        print("\n" + "="*80)
        
        # Visual bar chart by category
        print("\n📊 Visual Summary by Category:")
        print("-" * 80)
        for category, stats in summary["category_stats"].items():
            pass_rate = stats['pass_rate']
            bar_length = int(pass_rate * 40)
            bar = "█" * bar_length + "░" * (40 - bar_length)
            print(f"{category:<25} {bar} {pass_rate:>6.1%}")
        print("="*80)

print("✅ SummaryGenerator class defined")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 12. Main Execution

# COMMAND ----------

# MAGIC %md
# MAGIC ### Load Test Data

# COMMAND ----------

# Load your test data
# Expected columns: "query" (required), "user_id" (optional)
# Upload your Excel/CSV file with test queries

# OPTION 1: Load from uploaded file (uncomment and update filename)
test_data = pd.read_excel("chatbotv3_test_queries.xlsx")  # Update with your filename

# Add patient_name column (set to "there" as requested)
if 'patient_name' not in test_data.columns:
    test_data['patient_name'] = 'there'

# Ensure required columns exist
if 'query' not in test_data.columns:
    raise ValueError("Test data must have 'query' column")

# Set default user_id if not present
if 'user_id' not in test_data.columns:
    test_data['user_id'] = '12345'

print(f"✅ Loaded {len(test_data)} test queries")
print(f"Columns: {list(test_data.columns)}")
print("\nFirst few rows:")
print(test_data.head())

# OPTION 2: Sample data for testing (comment out if using uploaded file)
# test_data = pd.DataFrame({
#     "query": [
#         "Show me my glucose trends for this week",
#         "What was my average blood sugar yesterday?"
#     ],
#     "user_id": ["12345", "12345"],
#     "patient_name": ["there", "there"]
# })

# COMMAND ----------

# MAGIC %md
# MAGIC ### Initialize and Run Evaluation

# COMMAND ----------

# ========================================================================
# OPTION 1: Use Default Criteria (from Section 4)
# ========================================================================
# This uses the EVALUATION_CRITERIA and CRITERIA_CATEGORIES defined in Section 4
config = EvaluationConfig(
    config_source_dir=CONFIG_SOURCE_DIR,
    judge_model_name=JUDGE_MODEL_NAME,
    guidelines_path=GUIDELINES_PATH,
    relevant_criteria=EVALUATION_CRITERIA,
    criteria_categories=CRITERIA_CATEGORIES
)

# ========================================================================
# OPTION 2: Use Custom Criteria (Programmatic Approach)
# ========================================================================
# Uncomment and modify this section to specify custom criteria programmatically
# ========================================================================

# custom_criteria = [
#     "data_context_accuracy",
#     "tone_warmth_and_approachability",
#     "tone_professionalism",
#     "spelling_and_punctuation",
#     "medical_terminology_correctness"
# ]
# 
# custom_categories = {
#     "Accuracy": ["data_context_accuracy"],
#     "Tone & Style": [
#         "tone_warmth_and_approachability",
#         "tone_professionalism"
#     ],
#     "Language": [
#         "spelling_and_punctuation",
#         "medical_terminology_correctness"
#     ]
# }
# 
# # Create config with custom criteria
# config = create_custom_evaluation_config(
#     config_source_dir=CONFIG_SOURCE_DIR,
#     evaluation_criteria=custom_criteria,
#     criteria_categories=custom_categories,
#     judge_model_name=JUDGE_MODEL_NAME,
#     guidelines_path=GUIDELINES_PATH
# )

print("✅ Configuration initialized")
print(f"   Loaded {len(config.guideline_map)} guidelines")
print(f"   Labeled criteria: {len(config.labeled_criteria)}")
print(f"   Unlabeled criteria: {len(config.unlabeled_criteria)}")
print(f"   Total criteria to evaluate: {len(config.relevant_criteria)}")
print(f"   Categories: {len(config.criteria_categories)}")

# COMMAND ----------

# Initialize pipeline
pipeline = EvaluationPipeline(config)
print("✅ Evaluation pipeline initialized")

# COMMAND ----------

# Run evaluation
with mlflow.start_run(run_name=f"evaluation_{len(config.relevant_criteria)}_criteria_{datetime.now().strftime('%Y%m%d_%H%M%S')}"):
    results = pipeline.evaluate_dataset(test_data)
    
    # Log parameters
    mlflow.log_param("num_queries", len(test_data))
    mlflow.log_param("num_criteria", len(config.relevant_criteria))
    mlflow.log_param("judge_model", JUDGE_MODEL_NAME)
    mlflow.log_param("criteria_categories", len(config.criteria_categories))
    mlflow.log_param("criteria_list", ", ".join(config.relevant_criteria))
    
    print(f"✅ Evaluation complete: {len(results)} results")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 13. Generate Reports

# COMMAND ----------

# Generate timestamp for filenames
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

# Generate Excel report with embedded charts
num_criteria = len(config.relevant_criteria)
excel_filename = f"chatbot_evaluation_results_{num_criteria}_criteria_{timestamp}.xlsx"
excel_generator = ExcelReportGenerator(config.criteria_categories, config.relevant_criteria)
excel_generator.generate_report(results, excel_filename)

print(f"✅ Excel report generated: {excel_filename}")

# COMMAND ----------

# Generate summary statistics
summary_generator = SummaryGenerator(config.criteria_categories, config.relevant_criteria)
summary = summary_generator.generate_summary(results)
summary_generator.print_summary(summary)

# Save summary as JSON
summary_filename = f"evaluation_summary_{num_criteria}_criteria_{timestamp}.json"
with open(summary_filename, 'w') as f:
    json.dump(summary, f, indent=2)

print(f"\n✅ Summary saved: {summary_filename}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 14. Log Artifacts to MLflow

# COMMAND ----------

with mlflow.start_run(run_name=f"evaluation_{num_criteria}_criteria_{timestamp}"):
    # Log summary metrics
    mlflow.log_metric("overall_pass_rate", summary["overall_pass_rate"])
    mlflow.log_metric("total_queries", summary["total_queries"])
    mlflow.log_metric("queries_with_charts", summary["queries_with_charts"])
    
    # Log category pass rates
    for category, stats in summary["category_stats"].items():
        mlflow.log_metric(f"pass_rate_{category.lower().replace(' ', '_')}", stats["pass_rate"])
    
    # Log criterion pass rates
    for criterion, stats in summary["criterion_stats"].items():
        mlflow.log_metric(f"pass_rate_{criterion}", stats["pass_rate"])
    
    # Log artifacts
    mlflow.log_artifact(excel_filename)
    mlflow.log_artifact(summary_filename)
    
    print("✅ Artifacts logged to MLflow")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 15. Detailed Performance Analysis

# COMMAND ----------

def print_detailed_analysis(results: List[Dict[str, Any]], evaluation_criteria: List[str]):
    """Print detailed performance analysis."""
    
    print("\n" + "="*80)
    print("🔍 DETAILED PERFORMANCE ANALYSIS")
    print("="*80)
    
    # Analyze by route
    route_performance = {}
    for result in results:
        route = result.get("route", "unknown")
        if route not in route_performance:
            route_performance[route] = {"scores": [], "count": 0}
        
        # Calculate pass rate for this response
        passes = sum(1 for e in result["evaluations"].values() if e["assessment"] == "yes")
        total = len(result["evaluations"])
        pass_rate = passes / total if total > 0 else 0
        
        route_performance[route]["scores"].append(pass_rate)
        route_performance[route]["count"] += 1
    
    print("\n📊 Performance by Route:")
    print("-" * 80)
    print(f"{'Route':<20} {'Queries':<10} {'Avg Pass Rate':<15}")
    print("-" * 80)
    for route, data in sorted(route_performance.items()):
        avg_pass_rate = np.mean(data["scores"])
        print(f"{route:<20} {data['count']:<10} {avg_pass_rate:>6.1%}")
    
    # Find weakest criteria
    print("\n⚠️  Criteria Needing Improvement (Pass Rate < 80%):")
    print("-" * 80)
    
    criterion_performance = {}
    for criterion in evaluation_criteria:
        scores = []
        for result in results:
            eval_data = result["evaluations"].get(criterion, {})
            score = 1 if eval_data.get("assessment") == "yes" else 0
            scores.append(score)
        if scores:
            pass_rate = np.mean(scores)
            criterion_performance[criterion] = pass_rate
            if pass_rate < 0.8:
                criterion_name = criterion.replace('_', ' ').title()
                print(f"   {criterion_name:<40} {pass_rate:>6.1%}")
    
    # Find strongest criteria
    print("\n✅ Strongest Criteria (Pass Rate >= 95%):")
    print("-" * 80)
    for criterion, pass_rate in criterion_performance.items():
        if pass_rate >= 0.95:
            criterion_name = criterion.replace('_', ' ').title()
            print(f"   {criterion_name:<40} {pass_rate:>6.1%}")
    
    print("="*80)

print_detailed_analysis(results, config.relevant_criteria)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 16. Per-User Analysis (if applicable)

# COMMAND ----------

# Analysis by user_id if available
if results and any('user_id' in r for r in results):
    print("\n" + "="*80)
    print("👥 ANALYSIS BY USER")
    print("="*80)
    
    # Group by user_id
    user_results = {}
    for result in results:
        user_id = result.get('user_id', 'N/A')
        if user_id not in user_results:
            user_results[user_id] = []
        user_results[user_id].append(result)
    
    for user_id, user_data in user_results.items():
        print(f"\n👤 User ID: {user_id}")
        print(f"   Queries evaluated: {len(user_data)}")
        
        # Calculate average pass rate for this user
        all_assessments = []
        for result in user_data:
            for eval_data in result['evaluations'].values():
                all_assessments.append(1 if eval_data['assessment'] == 'yes' else 0)
        
        if all_assessments:
            user_avg = np.mean(all_assessments)
            print(f"   Average pass rate: {user_avg:.1%}")
        
        # Show routes used
        routes = [r.get('route', 'unknown') for r in user_data]
        route_counts = {r: routes.count(r) for r in set(routes)}
        print(f"   Routes: {dict(route_counts)}")
        
        # Show chart statistics
        user_charts = sum(1 for r in user_data if r.get('show_chart', False))
        print(f"   Responses with charts: {user_charts}/{len(user_data)}")
else:
    print("\n⚠️  No user_id information available for per-user analysis")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Notes & Documentation
# MAGIC
# MAGIC ### Key Features - Configurable Evaluation Criteria:
# MAGIC
# MAGIC 1. **Flexible Evaluation Criteria**:
# MAGIC    - **Default: 11 Quality Criteria** organized into 5 categories:
# MAGIC      - **Accuracy (1)**: data_context_accuracy
# MAGIC      - **Coaching & Content (1)**: coaching_appropriateness
# MAGIC      - **Tone & Style (4)**: tone_warmth_and_approachability, tone_professionalism, 
# MAGIC        reading_level_appropriateness, message_length_appropriateness
# MAGIC      - **Language & Terminology (4)**: spelling_and_punctuation, medical_terminology_correctness, 
# MAGIC        inappropriate_language_avoidance, language_variety_and_flow
# MAGIC      - **Format & Presentation (1)**: response_structure_and_format
# MAGIC    - **Customizable**: You can specify which criteria to evaluate and how to categorize them
# MAGIC
# MAGIC 2. **How to Customize Evaluation Criteria**:
# MAGIC    
# MAGIC    **Method 1 - Edit Section 4 (Simple)**:
# MAGIC    - Go to Section 4 "Define Evaluation Criteria"
# MAGIC    - Uncomment and modify `CUSTOM_EVALUATION_CRITERIA` and `CUSTOM_CRITERIA_CATEGORIES`
# MAGIC    - Set `USE_CUSTOM_CRITERIA = True`
# MAGIC    
# MAGIC    **Method 2 - Programmatic (Recommended for scripting)**:
# MAGIC    - In Section 11, uncomment the custom criteria configuration
# MAGIC    - Modify `custom_criteria` list and `custom_categories` dict
# MAGIC    - Use `create_custom_evaluation_config()` helper function
# MAGIC    
# MAGIC    **Example - Evaluate only 3 criteria**:
# MAGIC    ```python
# MAGIC    custom_criteria = [
# MAGIC        "data_context_accuracy",
# MAGIC        "tone_professionalism",
# MAGIC        "spelling_and_punctuation"
# MAGIC    ]
# MAGIC    
# MAGIC    custom_categories = {
# MAGIC        "Accuracy": ["data_context_accuracy"],
# MAGIC        "Quality": ["tone_professionalism", "spelling_and_punctuation"]
# MAGIC    }
# MAGIC    
# MAGIC    config = create_custom_evaluation_config(
# MAGIC        config_source_dir=CONFIG_SOURCE_DIR,
# MAGIC        evaluation_criteria=custom_criteria,
# MAGIC        criteria_categories=custom_categories
# MAGIC    )
# MAGIC    ```
# MAGIC
# MAGIC 3. **Data Context Accuracy Criterion**:
# MAGIC    - Verifies that numeric data from source data context is correctly transferred to bot response
# MAGIC    - Checks for hallucinated numbers, calculation errors, and data mismatches
# MAGIC    - Data context is passed to the LLM judge for verification
# MAGIC
# MAGIC 4. **Direct Bot Approach**:
# MAGIC    - Uses DirectBotManager to call bots directly
# MAGIC    - No ChatAgent wrapper overhead
# MAGIC    - Routes using router_bot.router_decision()
# MAGIC    - Invokes bot chains with bot.get_chain().invoke()
# MAGIC    - Supports `env` parameter for environment selection (uat/prod)
# MAGIC
# MAGIC 5. **Enhanced Reporting**:
# MAGIC    - Category-level performance metrics
# MAGIC    - Per-criterion detailed analysis
# MAGIC    - Visual summaries by category
# MAGIC    - Identification of weak and strong areas
# MAGIC    - Dynamic filenames based on number of criteria evaluated
# MAGIC
# MAGIC ### Benefits:
# MAGIC - ✅ **Fully configurable** - evaluate any subset of criteria you choose
# MAGIC - ✅ Data context accuracy verification (source data to response)
# MAGIC - ✅ Clear categorization of evaluation areas
# MAGIC - ✅ Faster evaluation without chart rendering (chart criteria disabled by default)
# MAGIC - ✅ Self-contained Excel file (easy to share)
# MAGIC - ✅ Direct bot control (no wrapper overhead)
# MAGIC - ✅ Category-level insights for targeted improvement
# MAGIC - ✅ Fully automated pipeline
# MAGIC - ✅ Validation checks for configuration consistency
# MAGIC
# MAGIC ### Available Criteria (from evaluation_guidelines.yaml):
# MAGIC All criteria must be defined in your `evaluation_guidelines.yaml` file. Common criteria include:
# MAGIC - `data_context_accuracy` - Numeric data accuracy
# MAGIC - `coaching_appropriateness` - Appropriate guidance
# MAGIC - `tone_warmth_and_approachability` - Friendly tone
# MAGIC - `tone_professionalism` - Professional tone
# MAGIC - `reading_level_appropriateness` - 5th-6th grade reading level
# MAGIC - `message_length_appropriateness` - Under 350 words
# MAGIC - `spelling_and_punctuation` - Correct spelling/punctuation
# MAGIC - `medical_terminology_correctness` - Proper medical terms
# MAGIC - `inappropriate_language_avoidance` - No stigmatizing language
# MAGIC - `language_variety_and_flow` - Natural language flow
# MAGIC - `response_structure_and_format` - Clear organization
# MAGIC
# MAGIC ### Output Files:
# MAGIC - `chatbot_evaluation_results_N_criteria_YYYYMMDD_HHMMSS.xlsx` - Excel report
# MAGIC - `evaluation_summary_N_criteria_YYYYMMDD_HHMMSS.json` - Summary statistics
# MAGIC   (where N = number of criteria you selected)